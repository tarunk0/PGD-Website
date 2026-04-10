from flask import Flask, render_template, jsonify, request, session, redirect, send_file, Response, make_response
import random
import os
import time
import json
import requests as http_requests
from datetime import datetime, timedelta
import pytz
from functools import wraps
from io import BytesIO, StringIO
from werkzeug.security import check_password_hash
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import Config
from models import db, Customer, Inquiry, Product, InquiryResponse, Subscriber

app = Flask(__name__)
app.config.from_object(Config)

# Initialize SQLAlchemy
db.init_app(app)

# Create database tables
with app.app_context():
    db.create_all()

# =================== CURRENCY CONVERSION CACHE ===================
CURRENCY_CACHE = {
    'rates': {
        'INR': 1.0,
        'USD': 0.01189,
        'EUR': 0.01094,
        'GBP': 0.00940,
        'AED': 0.04367,
        'AUD': 0.01838,
        'BDT': 1.4239,
        'BRL': 0.06778,
        'CAD': 0.01638,
        'CHF': 0.01049,
        'CNY': 0.08653,
        'EGP': 0.5915,
        'JPY': 1.828,
        'KWD': 0.003654,
        'MYR': 0.05274,
        'NGN': 19.07,
        'NZD': 0.02004,
        'QAR': 0.04329,
        'RUB': 1.1905,
        'SAR': 0.04459,
        'SGD': 0.01584,
        'THB': 0.4063,
        'TRY': 0.4598,
        'ZAR': 0.2158,
    },
    'last_updated': datetime(2000, 1, 1, tzinfo=pytz.UTC),
    'base_currency': 'INR'
}

def fetch_live_rates():
    """Fetch live exchange rates from open.er-api.com (free, no key required)"""
    try:
        url = 'https://open.er-api.com/v6/latest/INR'
        resp = http_requests.get(url, timeout=10, headers={'User-Agent': 'PGD-Website/1.0'})
        resp.raise_for_status()
        data = resp.json()
        if data.get('result') == 'success' and 'rates' in data:
            # Filter to only currencies we support
            supported = list(CURRENCY_CACHE['rates'].keys())
            new_rates = {k: v for k, v in data['rates'].items() if k in supported}
            if new_rates and 'INR' in new_rates:
                CURRENCY_CACHE['rates'] = new_rates
                CURRENCY_CACHE['last_updated'] = datetime.now(pytz.UTC)
                return True
    except Exception as e:
        print(f'Currency API fetch error: {e}')
    return False

# Fetch live rates at startup
fetch_live_rates()

# Authentication Helper Functions
def login_required(f):
    """Decorator to protect admin routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return redirect('/admin/login')
        return f(*args, **kwargs)
    return decorated_function

def check_admin_password(username, password):
    """Check if admin credentials are correct"""
    if username != app.config['ADMIN_USERNAME']:
        return False
    return check_password_hash(app.config['ADMIN_PASSWORD_HASH'], password)

def generate_pdf_report(inquiries):
    """Generate professional PDF report of inquiries with full details"""
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch, 
                           leftMargin=0.75*inch, rightMargin=0.75*inch)
    story = []
    
    # Custom styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#b87333'),
        spaceAfter=6,
        alignment=1,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.gray,
        spaceAfter=12,
        alignment=1,
        fontName='Helvetica'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#00172D'),
        spaceAfter=8,
        spaceBefore=8,
        fontName='Helvetica-Bold'
    )
    
    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#2d2d2d'),
        spaceAfter=6,
        leading=12
    )
    
    # Title and Date - using India timezone
    india_tz = pytz.timezone('Asia/Kolkata')
    generated_time = datetime.now(india_tz)
    story.append(Paragraph("📊 PGD - CUSTOMER INQUIRIES REPORT", title_style))
    story.append(Paragraph(f"Generated on {generated_time.strftime('%Y-%m-%d at %H:%M:%S %Z')}", subtitle_style))
    story.append(Spacer(1, 0.15*inch))
    
    # Summary Statistics Section
    total = len(inquiries)
    new_count = sum(1 for i in inquiries if i.status == 'new')
    replied_count = sum(1 for i in inquiries if i.status == 'replied')
    converted_count = sum(1 for i in inquiries if i.status == 'converted')
    
    summary_data = [
        ['Total Inquiries', 'New', 'Replied', 'Converted'],
        [str(total), str(new_count), str(replied_count), str(converted_count)]
    ]
    
    summary_table = Table(summary_data, colWidths=[1.8*inch, 1.4*inch, 1.4*inch, 1.4*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00172D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9F6EE')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1.5, colors.HexColor('#b87333')),
        ('TOPPADDING', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.25*inch))
    
    # Detailed Inquiries Section
    if inquiries:
        story.append(Paragraph("DETAILED INQUIRIES", heading_style))
        
        for idx, inq in enumerate(inquiries, 1):
            # Inquiry header
            status_color = {
                'new': '#FFC107',
                'replied': '#17A2B8',
                'converted': '#28A745'
            }.get(inq.status, '#6C757D')
            
            # Convert inquiry date to India timezone
            india_tz = pytz.timezone('Asia/Kolkata')
            inquiry_date_ist = inq.created_at.replace(tzinfo=pytz.UTC).astimezone(india_tz)
            
            header_text = f"<b>Inquiry #{inq.id}</b> | Status: <font color='{status_color}'><b>{inq.status.upper()}</b></font> | Date: {inquiry_date_ist.strftime('%Y-%m-%d %H:%M IST')}"
            story.append(Paragraph(header_text, body_style))
            
            # Customer Information
            customer_info = f"<b>Customer:</b> {inq.customer.name}<br/><b>Email:</b> {inq.customer.email}<br/><b>Phone:</b> {inq.customer.phone or 'N/A'}"
            story.append(Paragraph(customer_info, body_style))
            
            # Message Section
            story.append(Paragraph("<b>Inquiry Message:</b>", body_style))
            message_text = inq.message.replace('<', '&lt;').replace('>', '&gt;')
            story.append(Paragraph(message_text, body_style))
            
            # Additional details - display product interest
            if inq.product_id:
                product = Product.query.get(inq.product_id)
                if product:
                    product_text = f"<b>Product Interest:</b> {product.name} ({product.category.title()})"
                else:
                    product_text = f"<b>Product Interest:</b> {inq.product_category or 'Product'}"
            elif inq.product_category:
                product_text = f"<b>Product Interest:</b> {inq.product_category.title()}"
            else:
                product_text = "<b>Product Interest:</b> General enquiry"
            story.append(Paragraph(product_text, body_style))
            
            # Separator line
            story.append(Spacer(1, 0.05*inch))
            story.append(Paragraph("_" * 80, body_style))
            story.append(Spacer(1, 0.1*inch))
            
            # Page break after every 5 inquiries (optional, for long lists)
            if idx % 5 == 0 and idx < total:
                story.append(PageBreak())
    
    # Footer
    story.append(Spacer(1, 0.2*inch))
    footer_text = f"<i>This report was automatically generated by PGD Admin System. For official records only.</i>"
    story.append(Paragraph(footer_text, subtitle_style))
    
    doc.build(story)
    pdf_buffer.seek(0)
    return pdf_buffer

def generate_excel_report(inquiries):
    """Generate Excel report of inquiries"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiries"
    
    # Styling
    header_fill = PatternFill(start_color="00172D", end_color="00172D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['ID', 'Customer Name', 'Email', 'Phone', 'Message', 'Status', 'Product Interest', 'Date', 'Last Updated']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Color mapping for status
    status_colors = {
        'new': 'FFC107',
        'replied': '17A2B8',
        'converted': '28A745'
    }
    
    # Data rows
    for inquiry in inquiries:
        status_color = status_colors.get(inquiry.status, 'FFFFFF')
        
        # Get product interest - lookup product name if product_id exists, fallback to product_category
        product_interest = 'General enquiry'
        if inquiry.product_id:
            product = Product.query.get(inquiry.product_id)
            if product:
                product_interest = f"{product.name} ({product.category.title()})"
        elif inquiry.product_category and inquiry.product_category.strip():
            product_interest = inquiry.product_category.title()
        
        # Convert to India timezone
        india_tz = pytz.timezone('Asia/Kolkata')
        created_ist = inquiry.created_at.replace(tzinfo=pytz.UTC).astimezone(india_tz)
        updated_ist = inquiry.last_updated.replace(tzinfo=pytz.UTC).astimezone(india_tz)
        
        ws.append([
            inquiry.id,
            inquiry.customer.name,
            inquiry.customer.email,
            inquiry.customer.phone or '',
            inquiry.message,
            inquiry.status.upper(),
            product_interest,
            created_ist.strftime('%Y-%m-%d %H:%M:%S IST'),
            updated_ist.strftime('%Y-%m-%d %H:%M:%S IST')
        ])
        
        # Style data row
        current_row = ws.max_row
        for col_num, cell in enumerate(ws[current_row], 1):
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            if col_num == 6:  # Status column
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 30  # Product Interest (was Product ID)
    ws.column_dimensions['H'].width = 20  # Date
    ws.column_dimensions['I'].width = 20  # Last Updated
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

PRODUCTS = [
    {"id": 1, "name": "Pure Copper Bottle", "category": "copper", "price": "2199", "purity": "99.9%", "image": "copper-bottle", "desc": "Handcrafted pure copper water bottle, traditionally forged in India."},
    {"id": 2, "name": "Copper Jug Set", "category": "copper", "price": "4299", "purity": "99.9%", "image": "copper-jug", "desc": "Elegant copper jug with two tumblers, ideal for ayurvedic water storage."},
    {"id": 3, "name": "Copper Plate Set", "category": "copper", "price": "2999", "purity": "99.5%", "image": "copper-plate", "desc": "Traditional copper dining plates with hand-embossed motifs."},
    {"id": 4, "name": "Ceramic Vase – Indigo", "category": "ceramic", "price": "3499", "purity": "Grade A", "image": "ceramic-vase", "desc": "Hand-painted blue pottery ceramic vase from Jaipur artisans."},
    {"id": 5, "name": "Ceramic Bowl Set", "category": "ceramic", "price": "2599", "purity": "Grade A", "image": "ceramic-bowl", "desc": "Artisan glazed ceramic bowls, kiln-fired with traditional techniques."},
    {"id": 6, "name": "Terracotta Planter", "category": "handicraft", "price": "1799", "purity": "Handmade", "image": "terracotta", "desc": "Sun-dried terracotta planter with tribal geometric carvings."},
    {"id": 7, "name": "Brass Deity Figurine", "category": "handicraft", "price": "4799", "purity": "Handmade", "image": "brass", "desc": "Lost-wax cast brass figurine, centuries-old Dhokra craft tradition."},
    {"id": 8, "name": "Copper Hammered Mug", "category": "copper", "price": "1599", "purity": "99.9%", "image": "copper-mug", "desc": "Hand-hammered copper mug with a rustic, artisanal finish."},
]

DESTINATIONS = [
    {"city": "New York", "country": "USA", "lat": 40.71, "lng": -74.00, "shipments": 1240},
    {"city": "London", "country": "UK", "lat": 51.50, "lng": -0.12, "shipments": 980},
    {"city": "Toronto", "country": "Canada", "lat": 43.65, "lng": -79.38, "shipments": 760},
    {"city": "Mexico City", "country": "Mexico", "lat": 19.43, "lng": -99.13, "shipments": 430},
    {"city": "Sydney", "country": "Australia", "lat": -33.86, "lng": 151.20, "shipments": 310},
    {"city": "Dubai", "country": "UAE", "lat": 25.20, "lng": 55.27, "shipments": 520},
]

# Populate Product table from PRODUCTS list if empty
with app.app_context():
    if Product.query.count() == 0:
        for prod in PRODUCTS:
            product = Product(
                id=prod['id'],
                name=prod['name'],
                category=prod['category'],
                price=prod['price'],
                purity=prod['purity'],
                description=prod['desc'],
                image=prod['image']
            )
            db.session.add(product)
        db.session.commit()

# =================== AUTHENTICATION ROUTES ======================

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    """Admin login page and handler"""
    if request.method == "POST":
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if check_admin_password(username, password):
            session['admin_logged_in'] = True
            session['admin_username'] = username
            session.permanent = True
            app.permanent_session_lifetime = timedelta(hours=24)
            return jsonify({"success": True, "message": "Logged in successfully"}), 200
        else:
            return jsonify({"success": False, "message": "Invalid credentials"}), 401
    
    return render_template("admin_login.html")

@app.route("/admin/logout")
def admin_logout():
    """Logout admin user"""
    session.clear()
    return redirect("/admin/login")

@app.route("/admin/check-session")
def check_session():
    """Check if admin is logged in"""
    return jsonify({
        "authenticated": 'admin_logged_in' in session
    }), 200

@app.route("/admin/inquiries")
@login_required
def admin_inquiries():
    """Admin dashboard to view all inquiries"""
    return render_template("admin_inquiries.html")

@app.route("/api/inquiries/export/pdf")
@login_required
def export_inquiries_pdf():
    """Export inquiries as PDF"""
    try:
        inquiries = Inquiry.query.order_by(Inquiry.created_at.desc()).all()
        pdf_buffer = generate_pdf_report(inquiries)
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'PGD_Inquiries_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/inquiries/export/excel")
@login_required
def export_inquiries_excel():
    """Export inquiries as Excel"""
    try:
        inquiries = Inquiry.query.order_by(Inquiry.created_at.desc()).all()
        excel_buffer = generate_excel_report(inquiries)
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'PGD_Inquiries_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# =================== SEO ROUTES ===================

@app.route("/robots.txt")
def robots_txt():
    """Serve robots.txt for search engine crawlers"""
    content = """User-agent: *
Allow: /
Disallow: /admin/
Disallow: /api/

Sitemap: https://pgd-website.fly.dev/sitemap.xml
"""
    return Response(content, mimetype="text/plain")

@app.route("/sitemap.xml")
def sitemap_xml():
    """Auto-generate XML sitemap for search engines"""
    base_url = "https://pgd-website.fly.dev"
    now = datetime.now(pytz.UTC).strftime("%Y-%m-%d")

    pages = [
        {"loc": "/", "changefreq": "weekly", "priority": "1.0"},
        {"loc": "/about", "changefreq": "monthly", "priority": "0.8"},
        {"loc": "/catalog", "changefreq": "weekly", "priority": "0.9"},
    ]

    xml = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml += '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    for page in pages:
        xml += "  <url>\n"
        xml += f"    <loc>{base_url}{page['loc']}</loc>\n"
        xml += f"    <lastmod>{now}</lastmod>\n"
        xml += f"    <changefreq>{page['changefreq']}</changefreq>\n"
        xml += f"    <priority>{page['priority']}</priority>\n"
        xml += "  </url>\n"
    xml += "</urlset>"

    return Response(xml, mimetype="application/xml")

# =================== SECURITY HEADERS ===================

@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    return response

# =================== PUBLIC ROUTES ===================

@app.route("/")
def home():
    return render_template("index.html", products=PRODUCTS[:6], destinations=DESTINATIONS)

@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/catalog")
def catalog():
    return render_template("catalog.html", products=PRODUCTS)

@app.route("/api/shipment-count")
def shipment_count():
    base = 4827
    live = base + random.randint(0, 99)
    return jsonify({"count": live, "active_routes": 6})

@app.route("/api/currency-rates")
def get_currency_rates():
    """Get current currency conversion rates (cached, updates daily)"""
    try:
        # Check if cache is older than 24 hours
        cache_age = (datetime.now(pytz.UTC) - CURRENCY_CACHE['last_updated']).total_seconds()
        if cache_age > 86400:  # 24 hours in seconds
            fetch_live_rates()
        
        return jsonify({
            "success": True,
            "rates": CURRENCY_CACHE['rates'],
            "base": CURRENCY_CACHE['base_currency'],
            "timestamp": CURRENCY_CACHE['last_updated'].isoformat()
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/inquiries", methods=["POST"])
def create_inquiry():
    """Save customer inquiry to database"""
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data or not data.get('name') or not data.get('email'):
            return jsonify({"error": "Name and email are required"}), 400
        
        # Check if customer exists
        customer = Customer.query.filter_by(email=data.get('email')).first()
        if not customer:
            customer = Customer(
                name=data.get('name'),
                email=data.get('email'),
                phone=data.get('phone', ''),
                country=data.get('country', ''),
                company_name=data.get('company_name', '')
            )
            db.session.add(customer)
            db.session.flush()
        else:
            # Update phone/country/company if new data is provided and different
            updated = False
            if data.get('phone') and data.get('phone') != customer.phone:
                customer.phone = data.get('phone')
                updated = True
            if data.get('country') and data.get('country') != customer.country:
                customer.country = data.get('country')
                updated = True
            if data.get('company_name') and data.get('company_name') != customer.company_name:
                customer.company_name = data.get('company_name')
                updated = True
            if updated:
                db.session.add(customer)
                db.session.flush()
        
        # Create inquiry
        inquiry = Inquiry(
            customer_id=customer.id,
            product_id=data.get('product_id'),
            product_category=data.get('product_category', ''),
            message=data.get('message', ''),
            status='new'
        )
        db.session.add(inquiry)
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "Inquiry saved successfully",
            "inquiry_id": inquiry.id
        }), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route("/api/inquiries", methods=["GET"])
def get_inquiries():
    """Get all inquiries (admin view)"""
    try:
        inquiries = Inquiry.query.order_by(Inquiry.created_at.desc()).all()
        result = []
        
        for inquiry in inquiries:
            # Get product interest - lookup product name if product_id exists, fallback to product_category
            product_interest = '-'
            if inquiry.product_id:
                product = Product.query.get(inquiry.product_id)
                if product:
                    product_interest = f"{product.name} ({product.category.title()})"
            elif inquiry.product_category and inquiry.product_category.strip():
                product_interest = inquiry.product_category.title()
            else:
                product_interest = 'General enquiry'
            
            # Convert to India timezone (IST)
            india_tz = pytz.timezone('Asia/Kolkata')
            created_ist = inquiry.created_at.replace(tzinfo=pytz.UTC).astimezone(india_tz)
            updated_ist = inquiry.last_updated.replace(tzinfo=pytz.UTC).astimezone(india_tz)
            
            result.append({
                "id": inquiry.id,
                "customer_name": inquiry.customer.name,
                "customer_email": inquiry.customer.email,
                "customer_phone": inquiry.customer.phone,
                "message": inquiry.message,
                "status": inquiry.status,
                "product_id": inquiry.product_id,
                "product_interest": product_interest,
                "created_at": created_ist.strftime("%Y-%m-%d %H:%M:%S"),
                "last_updated": updated_ist.strftime("%Y-%m-%d %H:%M:%S")
            })
        
        return jsonify({
            "success": True,
            "total": len(result),
            "inquiries": result
        }), 200
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/inquiries/<int:inquiry_id>", methods=["PUT"])
def update_inquiry_status(inquiry_id):
    """Update inquiry status"""
    try:
        inquiry = Inquiry.query.get(inquiry_id)
        if not inquiry:
            return jsonify({"error": "Inquiry not found"}), 404
        
        data = request.get_json()
        if 'status' in data:
            inquiry.status = data['status']
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "Inquiry updated successfully"
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


# =================== NEWSLETTER / SUBSCRIBE ===================

def send_email(to_email, subject, html_body):
    """Send email via Resend API. Returns True on success."""
    api_key = app.config.get('RESEND_API_KEY', '')
    if not api_key:
        print(f"[EMAIL-SKIP] No RESEND_API_KEY set. Would send to {to_email}: {subject}")
        return False
    try:
        resp = http_requests.post(
            'https://api.resend.com/emails',
            headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
            json={
                'from': app.config.get('MAIL_FROM', 'PGD <onboarding@resend.dev>'),
                'to': [to_email],
                'subject': subject,
                'html': html_body,
            },
            timeout=10,
        )
        resp.raise_for_status()
        return True
    except Exception as e:
        print(f"[EMAIL-ERROR] {e}")
        return False


def build_confirm_email(confirm_url):
    """Build confirmation email HTML."""
    return f'''
    <div style="max-width:520px;margin:0 auto;font-family:'Montserrat',Arial,sans-serif;background:#F9F6EE;border:1px solid #e8e0d4;border-radius:4px;overflow:hidden;">
      <div style="background:#00172D;padding:28px 24px;text-align:center;border-bottom:3px solid #b87333;">
        <span style="font-family:'Cinzel',Georgia,serif;font-size:28px;color:#d4956b;letter-spacing:4px;">PGD</span>
      </div>
      <div style="padding:32px 24px;text-align:center;">
        <h2 style="font-family:'Cinzel',Georgia,serif;color:#00172D;margin:0 0 12px;">Confirm Your Subscription</h2>
        <p style="color:#444;font-size:14px;line-height:1.6;margin:0 0 24px;">
          Thank you for subscribing to PGD newsletters. Click below to confirm your email and start receiving updates on new products, offers, and export insights.
        </p>
        <a href="{confirm_url}" style="display:inline-block;background:#b87333;color:#fff;text-decoration:none;padding:12px 32px;border-radius:3px;font-size:13px;font-weight:600;letter-spacing:1px;text-transform:uppercase;">Confirm Subscription</a>
        <p style="color:#999;font-size:11px;margin-top:24px;">If you didn't subscribe, ignore this email.</p>
      </div>
      <div style="background:#00172D;padding:14px;text-align:center;">
        <span style="color:rgba(249,246,238,0.3);font-size:10px;letter-spacing:1px;">© 2026 Pandra Global Dynamics Pvt Ltd</span>
      </div>
    </div>'''


def build_newsletter_email(body_html, unsubscribe_url):
    """Wrap admin-composed content in PGD branded email template."""
    return f'''
    <div style="max-width:560px;margin:0 auto;font-family:'Montserrat',Arial,sans-serif;background:#F9F6EE;border:1px solid #e8e0d4;border-radius:4px;overflow:hidden;">
      <div style="background:#00172D;padding:28px 24px;text-align:center;border-bottom:3px solid #b87333;">
        <span style="font-family:'Cinzel',Georgia,serif;font-size:28px;color:#d4956b;letter-spacing:4px;">PGD</span>
        <div style="color:rgba(249,246,238,0.5);font-size:10px;letter-spacing:2px;margin-top:6px;text-transform:uppercase;">Pandra Global Dynamics</div>
      </div>
      <div style="padding:32px 28px;color:#333;font-size:14px;line-height:1.7;">
        {body_html}
      </div>
      <div style="background:#00172D;padding:16px 24px;text-align:center;border-top:1px solid rgba(184,115,51,0.2);">
        <span style="color:rgba(249,246,238,0.35);font-size:10px;letter-spacing:1px;">
          &copy; 2026 Pandra Global Dynamics Pvt Ltd &nbsp;&middot;&nbsp;
          <a href="{unsubscribe_url}" style="color:rgba(249,246,238,0.5);text-decoration:underline;">Unsubscribe</a>
        </span>
      </div>
    </div>'''


@app.route("/api/newsletter/send", methods=["POST"])
@login_required
def send_newsletter():
    """Send a newsletter to all confirmed active subscribers."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "Invalid request"}), 400

        subject = (data.get('subject') or '').strip()
        body_html = (data.get('body_html') or '').strip()

        if not subject:
            return jsonify({"success": False, "error": "Subject is required"}), 400
        if not body_html:
            return jsonify({"success": False, "error": "Email body is required"}), 400
        if len(subject) > 200:
            return jsonify({"success": False, "error": "Subject too long (max 200 chars)"}), 400

        subscribers = Subscriber.query.filter_by(is_active=True, is_confirmed=True).all()
        if not subscribers:
            return jsonify({"success": False, "error": "No confirmed subscribers to send to"}), 400

        site_url = app.config.get('SITE_URL', 'https://pgd-website.fly.dev')
        sent = 0
        failed = 0

        for subscriber in subscribers:
            unsubscribe_url = f"{site_url}/unsubscribe/{subscriber.unsubscribe_token}"
            full_html = build_newsletter_email(body_html, unsubscribe_url)
            ok = send_email(subscriber.email, subject, full_html)
            if ok:
                sent += 1
            else:
                failed += 1
            time.sleep(0.5)  # Respect Resend 2 req/sec rate limit

        return jsonify({
            "success": True,
            "message": f"Sent to {sent} subscriber{'s' if sent != 1 else ''}.",
            "sent": sent,
            "failed": failed,
            "total": len(subscribers),
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/subscribe", methods=["POST"])
def subscribe():
    """Subscribe to newsletter with double opt-in."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "Invalid request"}), 400

        email = (data.get('email') or '').strip().lower()
        name = (data.get('name') or '').strip()

        if not email or '@' not in email or '.' not in email.split('@')[-1]:
            return jsonify({"success": False, "error": "Please enter a valid email address"}), 400

        if len(email) > 255:
            return jsonify({"success": False, "error": "Email too long"}), 400

        existing = Subscriber.query.filter_by(email=email).first()
        if existing:
            if existing.is_active and existing.is_confirmed:
                return jsonify({"success": True, "message": "You're already subscribed!"}), 200
            if existing.is_active and not existing.is_confirmed:
                # Resend confirmation
                confirm_url = f"{app.config['SITE_URL']}/subscribe/confirm/{existing.confirm_token}"
                send_email(existing.email, "Confirm your PGD subscription", build_confirm_email(confirm_url))
                return jsonify({"success": True, "message": "Confirmation email resent. Please check your inbox."}), 200
            # Reactivate
            existing.is_active = True
            existing.is_confirmed = False
            existing.unsubscribed_at = None
            import secrets
            existing.confirm_token = secrets.token_urlsafe(32)
            existing.unsubscribe_token = secrets.token_urlsafe(32)
            if name:
                existing.name = name
            db.session.commit()
            confirm_url = f"{app.config['SITE_URL']}/subscribe/confirm/{existing.confirm_token}"
            send_email(existing.email, "Confirm your PGD subscription", build_confirm_email(confirm_url))
            return jsonify({"success": True, "message": "Welcome back! Please check your email to confirm."}), 200

        subscriber = Subscriber(email=email, name=name if name else None)
        db.session.add(subscriber)
        db.session.commit()

        confirm_url = f"{app.config['SITE_URL']}/subscribe/confirm/{subscriber.confirm_token}"
        send_email(email, "Confirm your PGD subscription", build_confirm_email(confirm_url))

        return jsonify({"success": True, "message": "Please check your email to confirm your subscription."}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/subscribe/confirm/<token>")
def confirm_subscription(token):
    """Confirm subscription via email link."""
    subscriber = Subscriber.query.filter_by(confirm_token=token, is_active=True).first()
    if not subscriber:
        return render_template("subscribe_status.html",
                               status="error",
                               message="Invalid or expired confirmation link.")
    if subscriber.is_confirmed:
        return render_template("subscribe_status.html",
                               status="already",
                               message="Your subscription is already confirmed!")
    subscriber.is_confirmed = True
    subscriber.confirmed_at = datetime.utcnow()
    db.session.commit()
    return render_template("subscribe_status.html",
                           status="success",
                           message="Your subscription is confirmed! You'll receive our latest updates.")


@app.route("/unsubscribe/<token>")
def unsubscribe(token):
    """Unsubscribe via email link (one-click)."""
    subscriber = Subscriber.query.filter_by(unsubscribe_token=token).first()
    if not subscriber:
        return render_template("subscribe_status.html",
                               status="error",
                               message="Invalid unsubscribe link.")
    if not subscriber.is_active:
        return render_template("subscribe_status.html",
                               status="already",
                               message="You've already been unsubscribed.")
    subscriber.is_active = False
    subscriber.unsubscribed_at = datetime.utcnow()
    db.session.commit()
    return render_template("subscribe_status.html",
                           status="unsubscribed",
                           message="You've been unsubscribed. We're sorry to see you go.")


# ---- Admin: Subscriber Management ----

@app.route("/admin/subscribers")
@login_required
def admin_subscribers():
    """Admin view of all subscribers."""
    return render_template("admin_subscribers.html")


@app.route("/api/subscribers", methods=["GET"])
@login_required
def get_subscribers():
    """Get all subscribers for admin."""
    try:
        subscribers = Subscriber.query.order_by(Subscriber.subscribed_at.desc()).all()
        return jsonify({
            "success": True,
            "subscribers": [{
                "id": s.id,
                "email": s.email,
                "name": s.name or "",
                "is_active": s.is_active,
                "is_confirmed": s.is_confirmed,
                "subscribed_at": s.subscribed_at.isoformat() if s.subscribed_at else None,
                "confirmed_at": s.confirmed_at.isoformat() if s.confirmed_at else None,
            } for s in subscribers],
            "total": len(subscribers),
            "active": sum(1 for s in subscribers if s.is_active and s.is_confirmed),
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/subscribers/export/csv")
@login_required
def export_subscribers_csv():
    """Export confirmed active subscribers as CSV."""
    subscribers = Subscriber.query.filter_by(is_active=True, is_confirmed=True).order_by(Subscriber.subscribed_at.desc()).all()
    output = StringIO()
    output.write("Email,Name,Subscribed At,Confirmed At\n")
    for s in subscribers:
        name = (s.name or "").replace('"', '""')
        output.write(f'"{s.email}","{name}","{s.subscribed_at}","{s.confirmed_at}"\n')
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=pgd_subscribers_{datetime.now().strftime("%Y%m%d")}.csv'}
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
